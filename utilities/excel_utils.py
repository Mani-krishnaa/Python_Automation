import openpyxl


# File -->WorkBok---Sheets__>Rows-->Cells


def get_row_count(file, sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.max_row


def get_column_count(file, sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.max_column


def read_data(file, sheetName, row_num, column_num):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.cell(row_num, column_num).value


def write_data(file, sheetName, row_num, column_num, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    sheet.cell(row_num, column_num).value = data
    workbook.save(file)
